from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
import os
import pandas as pd
import PyPDF2
import requests
import json
from werkzeug.utils import secure_filename
import tempfile
import csv
from io import StringIO, BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Load configuration
try:
    from config import DATABRICKS_CONFIG
    DATABRICKS_TOKEN = DATABRICKS_CONFIG['token']
    DATABRICKS_ENDPOINTS = DATABRICKS_CONFIG['endpoints']
except ImportError:
    print("Warning: config.py not found. Please copy config_template.py to config.py and update with your credentials.")
    DATABRICKS_TOKEN = 'YOUR_DATABRICKS_TOKEN_HERE'
    DATABRICKS_ENDPOINTS = {
        'claude-sonnet-4': 'https://dbc-3735add4-1cb6.cloud.databricks.com/serving-endpoints/databricks-claude-sonnet-4/invocations',
        'llama-3-70b': 'https://dbc-3735add4-1cb6.cloud.databricks.com/serving-endpoints/databricks-meta-llama-3-3-70b-instruct/invocations'
    }

# Output layouts available
OUTPUT_LAYOUTS = ['member', 'service_provider', 'bill_custom_detail']

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def parse_csv_data_dictionary(file_path):
    """Parse CSV data dictionary and extract table information"""
    try:
        df = pd.read_csv(file_path)
        return df.to_dict('records')
    except Exception as e:
        return {"error": f"Error parsing CSV: {str(e)}"}

def parse_pdf_data_dictionary(file_path):
    """Parse PDF data dictionary and extract text content"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return {"pdf_content": text}
    except Exception as e:
        return {"error": f"Error parsing PDF: {str(e)}"}

def load_output_layout(layout_name):
    """Load the specified output layout CSV file"""
    try:
        layout_path = f"output_layouts/{layout_name}.csv"
        df = pd.read_csv(layout_path)
        # Handle NaN values by replacing them with empty strings
        df = df.fillna('')
        return df.to_dict('records')
    except Exception as e:
        return {"error": f"Error loading output layout: {str(e)}"}

def filter_data_dictionary_by_tables(data_dict, table_names):
    """Filter data dictionary to include only specified table names"""
    if isinstance(data_dict, list):
        # For CSV data dictionary
        filtered_data = []
        for entry in data_dict:
            # Check if any column contains table name information
            for key, value in entry.items():
                if isinstance(value, str) and any(table_name.lower() in value.lower() for table_name in table_names):
                    filtered_data.append(entry)
                    break
        return filtered_data
    elif isinstance(data_dict, dict) and 'pdf_content' in data_dict:
        # For PDF data dictionary - extract relevant sections
        content = data_dict['pdf_content']
        relevant_sections = []
        lines = content.split('\n')
        for line in lines:
            if any(table_name.lower() in line.lower() for table_name in table_names):
                relevant_sections.append(line)
        return {"filtered_pdf_content": '\n'.join(relevant_sections)}
    return data_dict

def call_databricks_llm(endpoint_key, prompt, max_tokens=4000):
    """Call Databricks LLM endpoint with improved error handling and retry logic"""
    import time
    
    try:
        url = DATABRICKS_ENDPOINTS[endpoint_key]
        
        # Check if token is configured
        if DATABRICKS_TOKEN == 'YOUR_DATABRICKS_TOKEN_HERE':
            return {"error": "Databricks token not configured. Please update config.py with your actual token."}
        
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {DATABRICKS_TOKEN}'
        }
        
        payload = {
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "max_tokens": max_tokens,
            "temperature": 0.1
        }
        
        # Add timeout and retry logic
        max_retries = 3
        for attempt in range(max_retries):
            try:
                print(f"Attempting API call to {endpoint_key} (attempt {attempt + 1}/{max_retries})")
                
                response = requests.post(
                    url, 
                    headers=headers, 
                    json=payload,
                    timeout=(30, 120),  # (connection_timeout, read_timeout)
                    verify=True
                )
                
                if response.status_code == 200:
                    result = response.json()
                    print(f"API call successful")
                    return result
                elif response.status_code == 401:
                    return {"error": "Authentication failed. Please check your Databricks token in config.py"}
                elif response.status_code == 404:
                    return {"error": f"Endpoint not found. Please verify the endpoint URL: {url}"}
                elif response.status_code == 429:
                    if attempt < max_retries - 1:
                        wait_time = 2 ** attempt  # Exponential backoff
                        print(f"Rate limited. Waiting {wait_time} seconds before retry...")
                        time.sleep(wait_time)
                        continue
                    return {"error": "Rate limit exceeded. Please try again later."}
                else:
                    return {"error": f"API call failed with status {response.status_code}: {response.text}"}
                    
            except requests.exceptions.ConnectTimeout:
                if attempt < max_retries - 1:
                    print(f"Connection timeout. Retrying in 2 seconds...")
                    time.sleep(2)
                    continue
                return {"error": "Connection timeout. Please check your internet connection and try again."}
                
            except requests.exceptions.ReadTimeout:
                if attempt < max_retries - 1:
                    print(f"Read timeout. Retrying in 2 seconds...")
                    time.sleep(2)
                    continue
                return {"error": "Read timeout. The AI model is taking too long to respond. Please try again."}
                
            except requests.exceptions.ConnectionError as e:
                if attempt < max_retries - 1:
                    print(f"Connection error: {str(e)}. Retrying in 2 seconds...")
                    time.sleep(2)
                    continue
                return {"error": f"Connection error: Unable to connect to Databricks. Please check your internet connection and endpoint URL."}
        
        return {"error": "Max retries exceeded. Please try again later."}
            
    except Exception as e:
        return {"error": f"Unexpected error calling LLM: {str(e)}"}

def parse_mapping_result_to_structured_data(mapping_result):
    """Parse LLM mapping result into structured field mapping data"""
    try:
        # Extract the response content
        if isinstance(mapping_result, dict):
            if 'choices' in mapping_result and len(mapping_result['choices']) > 0:
                content = mapping_result['choices'][0].get('message', {}).get('content', '')
            elif 'predictions' in mapping_result and len(mapping_result['predictions']) > 0:
                content = mapping_result['predictions'][0].get('candidates', [{}])[0].get('content', '')
            elif 'error' in mapping_result:
                return {'error': mapping_result['error']}
            else:
                content = str(mapping_result)
        else:
            content = str(mapping_result)
        
        # Parse the structured field mappings
        field_mappings = []
        
        # Look for the structured format with TARGET_FIELD patterns
        lines = content.split('\n')
        current_field = {}
        
        for line in lines:
            line = line.strip()
            if not line or line == '---':
                if current_field and 'target_field' in current_field:
                    field_mappings.append(current_field)
                    current_field = {}
                continue
            
            # Parse structured field information
            if line.startswith('TARGET_FIELD:'):
                current_field['target_field'] = line.replace('TARGET_FIELD:', '').strip()
            elif line.startswith('TARGET_TYPE:'):
                current_field['target_type'] = line.replace('TARGET_TYPE:', '').strip()
            elif line.startswith('TARGET_DESC:'):
                current_field['target_desc'] = line.replace('TARGET_DESC:', '').strip()
            elif line.startswith('SOURCE_TABLE:'):
                current_field['source_table'] = line.replace('SOURCE_TABLE:', '').strip()
            elif line.startswith('SOURCE_COLUMN:'):
                current_field['source_column'] = line.replace('SOURCE_COLUMN:', '').strip()
            elif line.startswith('TRANSFORMATION:'):
                current_field['transformation'] = line.replace('TRANSFORMATION:', '').strip()
            elif line.startswith('CONFIDENCE:'):
                current_field['confidence'] = line.replace('CONFIDENCE:', '').strip()
            elif line.startswith('REASON:'):
                current_field['reason'] = line.replace('REASON:', '').strip()
        
        # Add the last field if it exists
        if current_field and 'target_field' in current_field:
            field_mappings.append(current_field)
        
        return {
            'full_content': content,
            'field_mappings': field_mappings
        }
        
    except Exception as e:
        return {'error': f'Error parsing mapping result: {str(e)}'}

def create_excel_mapping_report(mapping_data, layout_name, table_names, output_layout):
    """Create a clean Excel report with structured field mappings"""
    try:
        # Create a BytesIO object to store the Excel file
        excel_buffer = BytesIO()
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sheet 1: Summary
        summary_ws = wb.create_sheet("Summary")
        summary_data = [
            ["Healthcare Data Mapping Report"],
            [""],
            ["Generated Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Target Layout:", layout_name],
            ["Source Tables:", ", ".join(table_names)],
            ["Total Target Fields:", len(output_layout)],
            [""],
            ["Report Contents:"],
            ["• Field Mappings: Structured field-by-field mapping with confidence scores"],
            ["• AI-generated transformations with reasoning"],
            ["• Healthcare data compliance considerations"]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:  # Title
                    cell.font = Font(bold=True, size=16)
                elif row_num in [3, 4, 5, 6]:  # Summary info
                    if col_num == 1:
                        cell.font = Font(bold=True)
        
        # Sheet 2: Field Mappings (Main Sheet)
        mapping_ws = wb.create_sheet("Field Mappings")
        
        # Define headers as requested
        headers = [
            "Target Field",
            "Target Data Type", 
            "Target Description",
            "Source Table",
            "Source Column",
            "Transformation / Mapping Rules",
            "Confidence",
            "Reason (Why AI mapping the source field to target)"
        ]
        
        mapping_ws.append(headers)
        
        # Apply header style
        for cell in mapping_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Generate field mappings based on parsed AI content or output layout
        if 'field_mappings' in mapping_data and mapping_data['field_mappings']:
            # Use parsed AI field mappings
            for field_mapping in mapping_data['field_mappings']:
                row_data = [
                    field_mapping.get('target_field', 'N/A'),
                    field_mapping.get('target_type', 'VARCHAR(255)'),
                    field_mapping.get('target_desc', 'No description'),
                    field_mapping.get('source_table', 'TBD'),
                    field_mapping.get('source_column', 'TBD'),
                    field_mapping.get('transformation', 'Direct mapping'),
                    field_mapping.get('confidence', 'Medium'),
                    field_mapping.get('reason', 'AI analysis required')
                ]
                mapping_ws.append(row_data)
        
        elif output_layout:
            # Fallback: Generate mappings based on output layout
            ai_content = mapping_data.get('full_content', '')
            
            for idx, field in enumerate(output_layout):
                target_field = field.get('Field', f'field_{idx+1}')
                target_type = field.get('Type', 'VARCHAR(255)')
                target_desc = field.get('Description', 'Field description')
                
                # Extract mapping information from AI content for this field
                source_table = "TBD"
                source_column = "TBD" 
                transformation = "Direct mapping"
                confidence = "Medium"
                reason = "AI analysis required - please review the full content for detailed mapping logic"
                
                # Try to find relevant information in AI content
                if target_field.lower() in ai_content.lower():
                    # Look for common patterns in AI responses
                    for table in table_names:
                        if table.lower() in ai_content.lower():
                            source_table = table
                            break
                    
                    source_column = target_field  # Default assumption
                    confidence = "High" if any(keyword in ai_content.lower() for keyword in [target_field.lower(), 'map', 'transform']) else "Medium"
                    reason = f"Field identified in AI analysis. Common healthcare field mapping for {target_field}. Review full AI content for complete logic."
                    
                    # Check for transformation keywords
                    if any(keyword in ai_content.lower() for keyword in ['trim', 'cast', 'convert', 'case']):
                        transformation = "Data transformation required (see AI analysis)"
                    elif any(keyword in ai_content.lower() for keyword in ['join', 'lookup']):
                        transformation = "Multi-table join required"
                    else:
                        transformation = "Direct field mapping"
                
                # Add row to Excel
                row_data = [
                    target_field,
                    target_type,
                    target_desc,
                    source_table,
                    source_column,
                    transformation,
                    confidence,
                    reason
                ]
                
                mapping_ws.append(row_data)
        
        # Apply border styling to all data rows
        for row_num in range(2, mapping_ws.max_row + 1):
            for col_num in range(1, len(headers) + 1):
                cell = mapping_ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for column in mapping_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            mapping_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set specific column widths for better readability
        mapping_ws.column_dimensions['A'].width = 20  # Target Field
        mapping_ws.column_dimensions['B'].width = 15  # Target Data Type
        mapping_ws.column_dimensions['C'].width = 30  # Target Description
        mapping_ws.column_dimensions['D'].width = 20  # Source Table
        mapping_ws.column_dimensions['E'].width = 20  # Source Column
        mapping_ws.column_dimensions['F'].width = 40  # Transformation Rules
        mapping_ws.column_dimensions['G'].width = 12  # Confidence
        mapping_ws.column_dimensions['H'].width = 50  # Reason
        
        # Set row height for headers
        mapping_ws.row_dimensions[1].height = 30
        
        # Auto-adjust column widths for summary sheet
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        return None

def create_mapping_prompt(output_layout, data_dictionary, table_names):
    """Create a comprehensive prompt for the LLM to generate data mapping"""
    
    prompt = f"""
You are an expert US Healthcare Data Modeler and Data Analyst specializing in healthcare data standardization and integration. You have extensive experience working with major healthcare data sources including Cigna, Facets, and other US healthcare systems. Your expertise includes:

- US Healthcare data standards (FHIR, HL7, X12, SNOMED CT, ICD-10, CPT)
- Healthcare payer systems (Cigna, Facets, Epic, Cerner, etc.)
- Claims processing, member enrollment, provider networks, and billing systems
- Healthcare data compliance (HIPAA, HITECH)
- Complex data transformations and mappings for healthcare interoperability

**TASK**: Generate comprehensive SQL-like transformation logic to map multiple source healthcare data tables into standardized output layouts, ensuring data quality, compliance, and accurate healthcare data representation.

**TARGET OUTPUT LAYOUT**:
{json.dumps(output_layout, indent=2)}

**SOURCE DATA TABLES TO USE**:
{', '.join(table_names)}

**DATA DICTIONARY (Source Table Details)**:
{json.dumps(data_dictionary, indent=2)}

**REQUIREMENTS**:
1. **Data Aggregation & Integration**: Analyze and aggregate data from multiple small source tables to create comprehensive target layouts
2. **Healthcare-Specific Transformations**: Apply industry-standard transformations including:
   - CAST() for data type conversions (dates, amounts, codes)
   - Complex JOIN operations across multiple healthcare tables (members, claims, providers, benefits)
   - TRIM() and data cleansing for healthcare identifiers and codes
   - CASE/WHEN for healthcare business logic and coding transformations
   - COALESCE() and ISNULL() for handling missing healthcare data
   - String concatenation for composite healthcare identifiers
   - Date/time formatting for healthcare transaction dates
   - Healthcare code mapping (ICD-9 to ICD-10, procedure codes, etc.)

3. **Healthcare Data Standards Compliance**: 
   - Ensure HIPAA compliance in data handling
   - Apply FHIR and HL7 standards where applicable
   - Maintain healthcare data integrity and audit trails
   - Handle PHI (Protected Health Information) appropriately

4. **Source System Expertise**: 
   - Understand Cigna-specific data structures and business rules
   - Handle Facets system data models and transformations
   - Account for payer-specific claim processing logic
   - Map provider network and member enrollment data accurately

5. **Data Quality & Validation**:
   - Implement healthcare-specific data validation rules
   - Handle duplicate member records and claim adjustments
   - Validate healthcare amounts, dates, and code relationships
   - Ensure referential integrity across healthcare entities

**OUTPUT FORMAT**:
Provide a structured healthcare data mapping response with:

1. **FIELD_MAPPING_TABLE** (Primary Output):
For each target field, provide a detailed mapping in the following format:
```
TARGET_FIELD: [field_name]
TARGET_TYPE: [data_type]
TARGET_DESC: [description]
SOURCE_TABLE: [source_table_name]
SOURCE_COLUMN: [source_column_name]
TRANSFORMATION: [transformation_logic]
CONFIDENCE: [High/Medium/Low]
REASON: [detailed_explanation_with_data_dictionary_reference]
---
```

2. **SQL_TRANSFORMATION_QUERIES**: 
   - Complete SQL statements for key transformations
   - Multi-table JOIN operations with healthcare entity relationships
   - Aggregation logic for summarizing healthcare transactions

3. **IMPLEMENTATION_NOTES**:
   - Healthcare business rules applied
   - HIPAA compliance considerations
   - Data quality recommendations
   - Performance optimization for large healthcare datasets

**CRITICAL REQUIREMENTS**:
- For each target field, analyze the data dictionary to find the most appropriate source mapping
- Provide specific confidence levels based on field name similarity, data type compatibility, and business logic
- Reference specific data dictionary entries in your reasoning
- Consider healthcare industry standards and common field mappings
- Account for data transformations needed for healthcare compliance and standardization

Generate a comprehensive, structured field mapping that can be easily parsed and implemented for healthcare data integration.
"""
    
    return prompt

@app.route('/')
def index():
    return render_template('index.html', layouts=OUTPUT_LAYOUTS)

@app.route('/api/generate_mapping', methods=['POST'])
def generate_mapping():
    try:
        # Get form data
        selected_layout = request.form.get('layout')
        table_names = request.form.get('table_names', '').split(',')
        table_names = [name.strip() for name in table_names if name.strip()]
        llm_model = request.form.get('llm_model', 'claude-sonnet-4')
        
        # Validate inputs
        if not selected_layout or selected_layout not in OUTPUT_LAYOUTS:
            return jsonify({'error': 'Invalid or missing layout selection'}), 400
        
        if not table_names:
            return jsonify({'error': 'Please specify at least one source table name'}), 400
        
        # Handle data dictionary file upload
        data_dict_file = request.files.get('data_dictionary')
        if not data_dict_file or data_dict_file.filename == '':
            return jsonify({'error': 'Please upload a data dictionary file'}), 400
        
        # Save uploaded file
        filename = secure_filename(data_dict_file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        data_dict_file.save(file_path)
        
        # Parse data dictionary based on file type
        file_extension = filename.rsplit('.', 1)[1].lower()
        if file_extension == 'csv':
            data_dictionary = parse_csv_data_dictionary(file_path)
        elif file_extension == 'pdf':
            data_dictionary = parse_pdf_data_dictionary(file_path)
        else:
            return jsonify({'error': 'Unsupported file format. Please upload CSV or PDF.'}), 400
        
        # Check for parsing errors
        if isinstance(data_dictionary, dict) and 'error' in data_dictionary:
            return jsonify(data_dictionary), 400
        
        # Filter data dictionary by specified table names
        filtered_data_dict = filter_data_dictionary_by_tables(data_dictionary, table_names)
        
        # Load target output layout
        output_layout = load_output_layout(selected_layout)
        if isinstance(output_layout, dict) and 'error' in output_layout:
            return jsonify(output_layout), 400
        
        # Create prompt for LLM
        prompt = create_mapping_prompt(output_layout, filtered_data_dict, table_names)
        
        # Call Databricks LLM
        llm_response = call_databricks_llm(llm_model, prompt)
        
        # Parse mapping result for structured data
        parsed_mapping = parse_mapping_result_to_structured_data(llm_response)
        
        # Clean up uploaded file
        os.remove(file_path)
        
        # Prepare response
        response_data = {
            'layout': selected_layout,
            'table_names': table_names,
            'llm_model': llm_model,
            'mapping_result': llm_response,
            'parsed_mapping': parsed_mapping,
            'output_layout_fields': len(output_layout),
            'data_dict_entries': len(filtered_data_dict) if isinstance(filtered_data_dict, list) else 1,
            'excel_available': True
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@app.route('/api/layouts/<layout_name>')
def get_layout_preview(layout_name):
    """Get a preview of the selected output layout"""
    if layout_name not in OUTPUT_LAYOUTS:
        return jsonify({'error': 'Invalid layout name'}), 400
    
    layout_data = load_output_layout(layout_name)
    if isinstance(layout_data, dict) and 'error' in layout_data:
        return jsonify(layout_data), 400
    
    # Return first 10 fields for preview
    preview_data = layout_data[:10] if len(layout_data) > 10 else layout_data
    
    return jsonify({
        'layout_name': layout_name,
        'total_fields': len(layout_data),
        'preview': preview_data
    })

@app.route('/api/export_excel', methods=['POST'])
def export_mapping_to_excel():
    """Export mapping result to Excel file"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Extract required data
        mapping_result = data.get('mapping_result')
        layout_name = data.get('layout', 'unknown_layout')
        table_names = data.get('table_names', [])
        output_layout = data.get('output_layout', [])
        
        if not mapping_result:
            return jsonify({'error': 'No mapping result provided'}), 400
        
        # Parse the mapping result
        parsed_mapping = parse_mapping_result_to_structured_data(mapping_result)
        
        if 'error' in parsed_mapping:
            return jsonify({'error': parsed_mapping['error']}), 400
        
        # Create Excel file
        excel_buffer = create_excel_mapping_report(
            parsed_mapping, 
            layout_name, 
            table_names, 
            output_layout
        )
        
        if excel_buffer is None:
            return jsonify({'error': 'Failed to create Excel file'}), 500
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"healthcare_data_mapping_{layout_name}_{timestamp}.xlsx"
        
        # Return the Excel file
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error creating Excel export: {str(e)}'}), 500

@app.route('/api/test_connection/<model_name>')
def test_databricks_connection(model_name):
    """Test endpoint to verify Databricks LLM connection"""
    if model_name not in DATABRICKS_ENDPOINTS:
        return jsonify({'error': f'Invalid model name. Available models: {list(DATABRICKS_ENDPOINTS.keys())}'}), 400
    
    # Test with a simple prompt
    test_prompt = "Hello, please respond with 'Connection successful' if you can read this message."
    
    print(f"Testing connection to {model_name}...")
    result = call_databricks_llm(model_name, test_prompt, max_tokens=100)
    
    if 'error' in result:
        return jsonify({
            'model': model_name,
            'status': 'failed',
            'error': result['error'],
            'endpoint_url': DATABRICKS_ENDPOINTS[model_name]
        }), 500
    else:
        return jsonify({
            'model': model_name,
            'status': 'success',
            'response': result,
            'endpoint_url': DATABRICKS_ENDPOINTS[model_name]
        })

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
